import streamlit as st
import pandas as pd
from datetime import datetime
import openpyxl
from openpyxl import Workbook
from io import BytesIO

# Initialize session state
if 'uploaded_file' not in st.session_state:
    st.session_state.uploaded_file = None

# Set page configuration
st.set_page_config(page_title="VIBER BLAST UPLOADER", page_icon="📊", layout="wide")

# Custom CSS
st.markdown(
    """
    <style>
    .main-content {
        padding: 20px;
        background-color: #f5f5f5;
        border-radius: 10px;
        box-shadow: 0 4px 8px rgba(0, 0, 0, 0.1);
        color: #2b2b2b;
        margin-bottom: 20px;
    }
    .stButton > button {
        width: 100%;
        margin-bottom: 10px;
        background-color: #b0b0b0;
        color: #2b2b2b;
        border: none;
        border-radius: 5px;
        box-shadow: 0 2px 4px rgba(0, 0, 0, 0.2);
        transition: all 0.2s ease;
    }
    .stButton > button:hover {
        background-color: #909090;
        transform: scale(1.05);
        box-shadow: 0 4px 8px rgba(0, 0, 0, 0.3);
    }
    .stDownloadButton > button {
        background-color: #b0b0b0;
        color: #2b2b2b;
        border-radius: 5px;
        box-shadow: 0 2px 4px rgba(0, 0, 0, 0.2);
        transition: all 0.2s ease;
    }
    .stDownloadButton > button:hover {
        background-color: #909090;
        transform: scale(1.05);
        box-shadow: 0 4px 8px rgba(0, 0, 0, 0.3);
    }
    h1 {
        text-align: center;
        color: #2b2b2b;
    }
    .stDataFrame {
        border: 1px solid #d0d0d0;
        border-radius: 5px;
        background-color: #ffffff;
    }
    .footer {
        text-align: center;
        color: #666666;
        margin-top: 20px;
        font-size: 12px;
    }
    </style>
    """,
    unsafe_allow_html=True
)

# Main content
with st.container():
    st.markdown('<div class="main-content">', unsafe_allow_html=True)
    
    st.subheader("Viber Blast CSV Uploader")
    
    # File uploader
    uploaded_file = st.file_uploader("📤 Choose a CSV file", type=["csv"], key="viber_blast_uploader", help="Upload a CSV with columns: Client, Account No., Debtor Name, Contact No.")
    if uploaded_file is not None:
        st.session_state.uploaded_file = uploaded_file
        st.success("File uploaded successfully!")

    # Reset button
    if st.session_state.uploaded_file is not None:
        if st.button("🔄 Reset", help="Clear the uploaded file and reset"):
            st.session_state.uploaded_file = None
            st.rerun()

    # Sample data
    sample_data = {
        "Campaign": [""],
        "CH Code": [""],
        "First Name": [""],
        "Full Name": [""],
        "Last Name": [""],
        "Mobile Number": [""],
        "OB": [""]
    }
    sample_df = pd.DataFrame(sample_data)

    # Dynamic filename
    current_date = datetime.now().strftime("VIBER BLAST %b %d %Y %I:%M %p PST").upper()

    if st.session_state.uploaded_file is not None:
        try:
            # Read the CSV file
            df = pd.read_csv(st.session_state.uploaded_file)
            
            # Check for required columns
            required_columns = ["Client", "Account No.", "Debtor Name", "Contact No."]
            missing_columns = [col for col in required_columns if col not in df.columns]
            
            if missing_columns:
                st.error(f"The following required columns are missing: {', '.join(missing_columns)}")
            else:
                # Process Contact No. and Account No.
                df["Contact No."] = df["Contact No."].astype(str).str.strip().replace("nan", "")
                df["Account No."] = df["Account No."].astype(str).str.strip().replace("nan", "")
                
                # Filter out rows where Contact No. is not 11 digits
                initial_row_count = len(df)
                df = df[df["Contact No."].str.len() == 11]
                removed_rows = initial_row_count - len(df)
                if removed_rows > 0:
                    st.info(f"Removed {removed_rows} rows where Contact No. is not exactly 11 digits.")
                
                # Remove duplicates based on Account No.
                initial_row_count = len(df)
                df = df.drop_duplicates(subset=["Account No."], keep="first")
                if initial_row_count != len(df):
                    st.info(f"Removed {initial_row_count - len(df)} duplicate rows based on 'Account No.'.")
                
                # Check if any rows remain
                if len(df) == 0:
                    st.warning("No rows remain after filtering. Showing sample data only.")
                
                # Create summary table
                summary_df = pd.DataFrame({
                    "Campaign": df["Client"],
                    "CH Code": df["Account No."],
                    "First Name": [""] * len(df),
                    "Full Name": df["Debtor Name"],
                    "Last Name": [""] * len(df),
                    "Mobile Number": df["Contact No."],
                    "OB": [""] * len(df)
                })
                
                # Concatenate with sample data
                summary_df = pd.concat([summary_df, sample_df], ignore_index=True)
                
                # Display summary table
                st.subheader("Summary Table")
                st.dataframe(summary_df, use_container_width=True)
                
                # Create Excel file with openpyxl
                output = BytesIO()
                wb = Workbook()
                ws = wb.active
                ws.title = "Viber Blast"
                
                # Write headers
                headers = list(summary_df.columns)
                for col_num, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col_num).value = header
                
                # Write data
                for row_num, row in enumerate(summary_df.values, 2):
                    for col_num, value in enumerate(row, 1):
                        ws.cell(row=row_num, column=col_num).value = value
                        # Set Mobile Number and CH Code columns to text format
                        if headers[col_num-1] in ["Mobile Number", "CH Code"]:
                            ws.cell(row=row_num, column=col_num).number_format = '@'
                
                # Save to BytesIO
                wb.save(output)
                output.seek(0)
                
                # Download button for Excel file
                st.download_button(
                    label="📥 Download Summary Table as Excel",
                    data=output,
                    file_name=f"{current_date}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="download_summary"
                )
                
        except Exception as e:
            st.error(f"An error occurred while processing the file: {str(e)}")
    else:
        # Display sample data if no file is uploaded
        st.subheader("Sample Summary Table")
        st.dataframe(sample_df, use_container_width=True)
        # Download button for sample data (Excel)
        output = BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "Viber Blast"
        headers = list(sample_df.columns)
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num).value = header
        for row_num, row in enumerate(sample_df.values, 2):
            for col_num, value in enumerate(row, 1):
                ws.cell(row=row_num, column=col_num).value = value
                if headers[col_num-1] in ["Mobile Number", "CH Code"]:
                    ws.cell(row=row_num, column=col_num).number_format = '@'
        wb.save(output)
        output.seek(0)
        st.download_button(
            label="📥 Download",
            data=output,
            file_name=f"{current_date}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="download_sample"
        )
        st.info("Please upload a CSV file to generate the summary table with your data.")
    
    st.markdown('</div>', unsafe_allow_html=True)

# Footer
st.markdown(f'<div class="footer">Viber Blast Uploader | Version 1.0 | {datetime.now().strftime("%b %d, %Y %I:%M %p PST")}</div>', unsafe_allow_html=True)